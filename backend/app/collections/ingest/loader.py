"""Workbook -> CanonicalDataset, via openpyxl reading typed cells into Decimal/date.

Never via anydoc (see docs/ARCHITECTURE.md) — a Markdown table cell is a
formatted string that has already lost precision and type; anydoc is for LLM
context only, never a calculation source.

Business-rule exclusions (unknown customer reference, missing due date,
Cancelled status, non-positive amounts, etc.) are NOT filtered here. Every
row that can be structurally parsed is loaded as-is; filtering happens in
`calculate/`. Only a row missing its own primary identifier is skipped, and
that is recorded in `CanonicalDataset.load_warnings`, never silent.
"""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from app.collections.contracts import (
    CanonicalCustomer,
    CanonicalDataset,
    CanonicalInvoice,
    CanonicalPayment,
    RegionMap,
)
from app.collections.ingest.resolver import (
    CUSTOMER_FIELDS,
    INVOICE_FIELDS,
    PAYMENT_FIELDS,
    REGION_MAP_FIELDS,
    ColumnResolutionError,
    resolve_columns,
)

REQUIRED_SHEETS = ("Customers", "Invoices", "Payments", "Region_Mapping")


def _get_sheet(workbook: openpyxl.Workbook, name: str) -> Worksheet:
    if name not in workbook.sheetnames:
        raise ColumnResolutionError(name, ["<sheet not found in workbook>"])
    return workbook[name]


def _str(cell: object) -> str:
    return str(cell).strip()


def _optional_str(cell: object) -> str | None:
    if cell is None:
        return None
    text = str(cell).strip()
    return text or None


def _decimal(cell: object) -> Decimal:
    try:
        return Decimal(str(cell))
    except (InvalidOperation, TypeError) as exc:
        raise ValueError(f"Cannot parse decimal from {cell!r}") from exc


def _date(cell: object) -> date:
    if isinstance(cell, datetime):
        return cell.date()
    if isinstance(cell, date):
        return cell
    raise ValueError(f"Cannot parse date from {cell!r}")


def _optional_date(cell: object) -> date | None:
    if cell is None:
        return None
    return _date(cell)


def _bool_flag(cell: object) -> bool:
    text = _str(cell).strip().lower()
    return text in ("y", "yes", "true", "1", "active")


def _is_blank_row(row: tuple[object, ...]) -> bool:
    return all(cell is None for cell in row)


def _load_customers(sheet: Worksheet, warnings: list[str]) -> list[CanonicalCustomer]:
    header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    columns = resolve_columns("Customers", header, CUSTOMER_FIELDS)
    customers: list[CanonicalCustomer] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if _is_blank_row(row):
            continue
        customer_id = _optional_str(row[columns["customer_id"]])
        if customer_id is None:
            warnings.append("Customers: skipped a row with a blank Customer_ID")
            continue
        customers.append(
            CanonicalCustomer(
                customer_id=customer_id,
                customer_name=_str(row[columns["customer_name"]]),
                city=_str(row[columns["city"]]),
                state=_str(row[columns["state"]]),
                region=_optional_str(row[columns["region"]]),
                gstin=_optional_str(row[columns["gstin"]]),
                credit_limit=_decimal(row[columns["credit_limit"]]),
                active_flag=_bool_flag(row[columns["active_flag"]]),
            )
        )
    return customers


def _load_invoices(sheet: Worksheet, warnings: list[str]) -> list[CanonicalInvoice]:
    header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    columns = resolve_columns("Invoices", header, INVOICE_FIELDS)
    invoices: list[CanonicalInvoice] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if _is_blank_row(row):
            continue
        invoice_id = _optional_str(row[columns["invoice_id"]])
        if invoice_id is None:
            warnings.append("Invoices: skipped a row with a blank Invoice_ID")
            continue
        invoices.append(
            CanonicalInvoice(
                invoice_id=invoice_id,
                customer_id=_str(row[columns["customer_id"]]),
                invoice_date=_date(row[columns["invoice_date"]]),
                due_date=_optional_date(row[columns["due_date"]]),
                invoice_amount=_decimal(row[columns["invoice_amount"]]),
                tax_amount=_decimal(row[columns["tax_amount"]]),
                status=_str(row[columns["status"]]),
                salesperson=_optional_str(row[columns["salesperson"]]),
                source_system_ref=_optional_str(row[columns["source_system_ref"]]),
            )
        )
    return invoices


def _load_payments(sheet: Worksheet, warnings: list[str]) -> list[CanonicalPayment]:
    header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    columns = resolve_columns("Payments", header, PAYMENT_FIELDS)
    payments: list[CanonicalPayment] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if _is_blank_row(row):
            continue
        payment_id = _optional_str(row[columns["payment_id"]])
        if payment_id is None:
            warnings.append("Payments: skipped a row with a blank Payment_ID")
            continue
        payments.append(
            CanonicalPayment(
                payment_id=payment_id,
                customer_id=_str(row[columns["customer_id"]]),
                invoice_id=_str(row[columns["invoice_id"]]),
                payment_date=_date(row[columns["payment_date"]]),
                payment_amount=_decimal(row[columns["payment_amount"]]),
                payment_mode=_optional_str(row[columns["payment_mode"]]),
                bank_ref=_optional_str(row[columns["bank_ref"]]),
            )
        )
    return payments


def _load_region_map(sheet: Worksheet, warnings: list[str]) -> list[RegionMap]:
    header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    columns = resolve_columns("Region_Mapping", header, REGION_MAP_FIELDS)
    region_map: list[RegionMap] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if _is_blank_row(row):
            continue
        state = _optional_str(row[columns["state"]])
        if state is None:
            warnings.append("Region_Mapping: skipped a row with a blank State")
            continue
        region_map.append(RegionMap(state=state, region=_str(row[columns["region"]])))
    return region_map


def load_workbook(path: Path) -> CanonicalDataset:
    """Load the four required sheets into a CanonicalDataset.

    Raises ColumnResolutionError if a required sheet or column is missing.
    """
    workbook = openpyxl.load_workbook(path, data_only=True)
    warnings: list[str] = []
    customers = _load_customers(_get_sheet(workbook, "Customers"), warnings)
    invoices = _load_invoices(_get_sheet(workbook, "Invoices"), warnings)
    payments = _load_payments(_get_sheet(workbook, "Payments"), warnings)
    region_map = _load_region_map(_get_sheet(workbook, "Region_Mapping"), warnings)
    return CanonicalDataset(
        customers=customers,
        invoices=invoices,
        payments=payments,
        region_map=region_map,
        load_warnings=warnings,
    )
