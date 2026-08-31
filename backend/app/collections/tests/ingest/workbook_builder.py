"""Minimal in-memory workbook builder for service.py failure-path tests.

Deliberately separate from `scripts/make_fixtures.py` (Phase 9's polished
dataset_b/c/d fixtures for the presentation demo) -- this builds the
smallest possible workbook that exercises one specific load-stage failure,
not a realistic dataset.
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl

from app.collections.ingest.resolver import (
    CUSTOMER_FIELDS,
    INVOICE_FIELDS,
    PAYMENT_FIELDS,
    REGION_MAP_FIELDS,
)

SHEET_HEADERS = {
    "Customers": list(CUSTOMER_FIELDS.values()),
    "Invoices": list(INVOICE_FIELDS.values()),
    "Payments": list(PAYMENT_FIELDS.values()),
    "Region_Mapping": list(REGION_MAP_FIELDS.values()),
}

VALID_ROWS: dict[str, list[list[object]]] = {
    "Customers": [
        [
            "C001",
            "Acme Co",
            "Mumbai",
            "Maharashtra",
            "West",
            "27ABCDE1234F1Z5",
            100000,
            "Yes",
        ],
    ],
    "Invoices": [
        [
            "INV-1001",
            "C001",
            date(2026, 1, 1),
            date(2026, 1, 31),
            1000,
            0,
            "Approved",
            "Raj",
            "SRC-1",
        ],
    ],
    "Payments": [],
    "Region_Mapping": [
        ["Maharashtra", "West"],
    ],
}


def build_minimal_workbook(
    path: Path,
    omit_sheets: set[str] | None = None,
    omit_headers: dict[str, set[str]] | None = None,
    cell_overrides: dict[str, dict[tuple[int, int], object]] | None = None,
) -> None:
    """Write a minimal 4-sheet workbook to `path`, one row per data sheet.

    `omit_sheets`: sheet names to leave out entirely (SheetMissingError).
    `omit_headers`: {sheet_name: {header_text, ...}} columns to drop
        (ColumnResolutionError).
    `cell_overrides`: {sheet_name: {(row, col): value}} 1-indexed,
        row 1 is the header row -- used to poison a data cell
        (ValueError from the loader's `_decimal`/`_date` helpers).
    """
    omit_sheets = omit_sheets or set()
    omit_headers = omit_headers or {}
    cell_overrides = cell_overrides or {}

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    for sheet_name, headers in SHEET_HEADERS.items():
        if sheet_name in omit_sheets:
            continue
        sheet = workbook.create_sheet(sheet_name)
        dropped = omit_headers.get(sheet_name, set())
        kept_headers = [h for h in headers if h not in dropped]
        sheet.append(kept_headers)
        for row in VALID_ROWS[sheet_name]:
            kept_row = [
                value
                for header, value in zip(headers, row, strict=True)
                if header not in dropped
            ]
            sheet.append(kept_row)
        for (row_idx, col_idx), value in cell_overrides.get(sheet_name, {}).items():
            sheet.cell(row=row_idx, column=col_idx, value=value)

    workbook.save(path)
