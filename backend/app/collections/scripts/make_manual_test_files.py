"""Generates a local, gitignored folder of .xlsx files covering the
manual-testing cases a person would want when poking at the app by
hand (Postman or the /collections/upload page): clean pass, blocked
with many exceptions, missing sheet, missing column, corrupt file.

Run: cd backend && uv run python -m app.collections.scripts.make_manual_test_files
Output: <repo root>/test-files/ (gitignored, not committed)
"""

from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl

FIXTURES_DIR = Path(__file__).parent.parent / "fixtures"
REPO_ROOT = Path(__file__).parents[4]
OUT_DIR = REPO_ROOT / "test-files"


def main() -> None:
    OUT_DIR.mkdir(exist_ok=True)

    shutil.copy(FIXTURES_DIR / "dataset_a_original.xlsx", OUT_DIR / "01_blocked_many_exceptions.xlsx")
    shutil.copy(FIXTURES_DIR / "dataset_b_clean.xlsx", OUT_DIR / "02_passes_clean.xlsx")

    # Missing sheet: drop Region_Mapping entirely.
    wb = openpyxl.load_workbook(FIXTURES_DIR / "dataset_a_original.xlsx")
    del wb["Region_Mapping"]
    wb.save(OUT_DIR / "03_missing_sheet.xlsx")

    # Missing column: rename Invoices' Due_Date header so it's not found.
    wb = openpyxl.load_workbook(FIXTURES_DIR / "dataset_a_original.xlsx")
    sheet = wb["Invoices"]
    for cell in sheet[1]:
        if cell.value == "Due_Date":
            cell.value = "Due_Date_Renamed"
    wb.save(OUT_DIR / "04_missing_column.xlsx")

    # Corrupt file: not a real xlsx at all.
    (OUT_DIR / "05_corrupt.xlsx").write_bytes(b"this is not a valid xlsx file")

    # Empty workbook: correct sheets/headers, zero data rows.
    wb = openpyxl.load_workbook(FIXTURES_DIR / "dataset_a_original.xlsx")
    for name in wb.sheetnames:
        sheet = wb[name]
        for row in list(sheet.iter_rows(min_row=2)):
            sheet.delete_rows(row[0].row, 1)
    wb.save(OUT_DIR / "06_empty_no_data_rows.xlsx")

    print(f"Wrote 6 test files to {OUT_DIR}")


if __name__ == "__main__":
    main()
