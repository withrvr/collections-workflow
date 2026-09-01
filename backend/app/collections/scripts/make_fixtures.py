"""Generates dataset_b_clean.xlsx from dataset_a_original.xlsx: same
headers, same shape, but every row touched by any of the 14 exception
rules removed -- so the control gate passes on it by construction.

Run: cd backend && uv run python -m app.collections.scripts.make_fixtures

MASTER_PLAN.md section 3 names dataset_c_renamed.xlsx (Phase 10, renamed
columns) and dataset_d_broken.xlsx (Phase 10, corrupt/missing-sheet) as
the other two presentation fixtures -- not built here, since nothing in
Phase 5-8 needs them yet; this script grows to cover them when Phase 10
does.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl

from app.collections.config import settings
from app.collections.ingest.loader import REQUIRED_SHEETS, load_workbook
from app.collections.validate.engine import run_all_rules

SOURCE_PATH = Path(__file__).parent.parent / "fixtures" / "dataset_a_original.xlsx"
DATASET_B_PATH = Path(__file__).parent.parent / "fixtures" / "dataset_b_clean.xlsx"


def _tainted_ids(source_path: Path) -> tuple[set[str], set[str], set[str]]:
    """Natural keys touched by any exception row, plus every invoice
    belonging to a tainted customer and every payment against a tainted
    invoice -- so the remaining rows are internally consistent, not just
    individually rule-clean."""
    dataset = load_workbook(source_path)
    exception_rows = run_all_rules(dataset, settings.REPORT_DATE)

    tainted_customers = {r.customer_id for r in exception_rows if r.customer_id}
    tainted_invoices = {r.invoice_id for r in exception_rows if r.invoice_id}
    tainted_payments = {r.payment_id for r in exception_rows if r.payment_id}

    tainted_invoices |= {
        i.invoice_id for i in dataset.invoices if i.customer_id in tainted_customers
    }
    tainted_payments |= {
        p.payment_id for p in dataset.payments if p.invoice_id in tainted_invoices
    }
    tainted_payments |= {
        p.payment_id for p in dataset.payments if p.customer_id in tainted_customers
    }

    return tainted_customers, tainted_invoices, tainted_payments


def make_dataset_b(
    source_path: Path = SOURCE_PATH, dest_path: Path = DATASET_B_PATH
) -> None:
    tainted_customers, tainted_invoices, tainted_payments = _tainted_ids(source_path)

    source = openpyxl.load_workbook(source_path, data_only=True)
    dest = openpyxl.Workbook()
    dest.remove(dest.active)

    key_column = {"Customers": 0, "Invoices": 0, "Payments": 0, "Region_Mapping": None}
    tainted_by_sheet = {
        "Customers": tainted_customers,
        "Invoices": tainted_invoices,
        "Payments": tainted_payments,
        "Region_Mapping": set(),
    }

    for sheet_name in REQUIRED_SHEETS:
        source_sheet = source[sheet_name]
        dest_sheet = dest.create_sheet(sheet_name)
        rows = list(source_sheet.iter_rows(values_only=True))
        dest_sheet.append(rows[0])  # header

        tainted_ids = tainted_by_sheet[sheet_name]
        key_col = key_column[sheet_name]
        for row in rows[1:]:
            if (
                key_col is not None
                and row[key_col] is not None
                and str(row[key_col]) in tainted_ids
            ):
                continue
            dest_sheet.append(row)

    dest.save(dest_path)


def main() -> None:
    make_dataset_b()
    dataset = load_workbook(DATASET_B_PATH)
    exception_rows = run_all_rules(dataset, settings.REPORT_DATE)
    print(
        f"Wrote {DATASET_B_PATH.name}: {len(dataset.invoices)} invoices, {len(exception_rows)} exception(s)."
    )


if __name__ == "__main__":
    main()
